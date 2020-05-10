import sys

import openpyxl
from PyQt5 import QtCore
from PyQt5.QtCore import QSize
from PyQt5.QtGui import QIcon, QKeySequence, QPixmap
from PyQt5.QtWidgets import QMainWindow, QApplication, QAction, QStatusBar, QToolBar, QWidget, QTableWidget, \
    QVBoxLayout, QAbstractItemView, QHeaderView, QShortcut, QLabel, QHBoxLayout, QLayout, QMessageBox
from Core.Base import *
from Core.FileIO.XpaNew import MakeXpaNew
from GUI.CenteralWidget import MyCenteralWidget
from GUI.Dialogs import XpaNewActionDialog
from GUI.MessageBox import CriticalMsgBox, QuestionMsgBox
# from PIL import Image

NOT_SAVE = True


class TableWidgetPixmap(QPixmap):
    def __init__(self, imgpath):
        super().__init__()
        self.load(imgpath)


class MyTabBarList(QWidget):
    def __init__(self):
        super().__init__()
        self.TableWidget = MyTableWidgetInTabBar()
        # self.TableWidget.setEnabled(True)
        vbox = QVBoxLayout()
        vbox.addWidget(self.TableWidget)
        self.setLayout(vbox)


class MyTableWidgetInTabBar(QTableWidget):
    def __init__(self):
        super().__init__(0, 3)
        self.setAcceptDrops(True)
        self.setSelectionMode(QAbstractItemView.SingleSelection)
        # self.setDragDropMode(QAbstractItemView.InternalMove)
        # self.setDefaultDropAction(QtCore.Qt.CopyAction)
        # 테이블위젯의 헤더 크기를 조정
        self.setAlternatingRowColors(True)
        self.setFocusPolicy(QtCore.Qt.StrongFocus)
        self.setMinimumSize(QtCore.QSize(990, 630))
        self.verticalHeader().setDefaultSectionSize(170)
        self.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.verticalHeader().setSectionResizeMode(QHeaderView.Fixed)
        self.setHorizontalHeaderLabels(['작업 전', '작업 후', '전주 번호'])
        self.setVerticalScrollMode(QAbstractItemView.ScrollPerPixel)

        copyShortcut = QShortcut(QKeySequence.Copy, self)
        pasteShortcut = QShortcut(QKeySequence.Paste, self)

        copyShortcut.activated.connect(self.copy)
        pasteShortcut.activated.connect(self.paste)

    def scaleup(self):
        pass

    def scaledown(self):
        pass

    def copy(self):
        selectedRangeList = self.selectedRanges()
        if not selectedRangeList:
            return

        r = self.currentRow()
        c = self.currentColumn()
        selectedPoint = self.cellWidget(r, c)
        image = selectedPoint.imgpath

        QApplication.clipboard().setText(image)

    def paste(self):
        image = QApplication.clipboard().text()
        r = self.currentRow()
        c = self.currentColumn()
        widget = TableItem(image, self)
        self.setCellWidget(r, c, widget)

    def dragEnterEvent(self, e):
        if e.mimeData().hasUrls:
            e.accept()
        else:
            e.ignore()

    def dragMoveEvent(self, e):
        if e.mimeData().hasUrls:
            e.accept()
        else:
            e.ignore()

    def dropEvent(self, e):
        global NOT_SAVE
        ex = ['JPG', 'jpg', 'png', 'jpge']
        e.setDropAction(QtCore.Qt.CopyAction)
        e.accept()
        if e.mimeData().hasUrls():
            row = self.currentRow()
            col = self.currentColumn()
            if self.cellWidget(row, col):
                # 이미 사진이 들어가있는 경우, 사진을 지우고 새로운 사진을 넣는다.
                pass
            else:
                print('없음')
            url = e.mimeData().urls()[0]
            url = str(url.toLocalFile())
            if url.split('.')[-1] in ex:
                print('위젯 전')
                widget = TableItem(url, self)
                print('위젯 후')
                self.setCellWidget(row, col, widget)
                NOT_SAVE = True

    # def resize_image(self):



class TableItem(QWidget):
    def __init__(self, imgpath, pars, pixmap=None):
        super().__init__()
        self.imgpath = imgpath
        self.lbl = QLabel()
        self.img = TableWidgetPixmap(imgpath) if pixmap is None else pixmap
        self.img = self.img.scaled(pars.width() // 3 - 30, 140 + 10)
        self.lbl.setPixmap(self.img)
        self.widget_layout = QHBoxLayout()
        self.widget_layout.addWidget(self.lbl)
        self.widget_layout.setSizeConstraint(QLayout.SetFixedSize)
        self.setLayout(self.widget_layout)


class MainWindow(QMainWindow):
    wb = None

    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setDefault()
        self.MyStatusBar()
        self.createMenuBar()

    # 기본 설정들을 불러옴
    def setDefault(self):
        self.setWindowTitle('엑셀 사진')
        myCenteralWidget = MyCenteralWidget()
        self.setCentralWidget(myCenteralWidget)

    def createMenuBar(self):
        # 파일 #
        # 메뉴 생성
        file_menu = self.menuBar().addMenu("&File")

        # 툴바 생성
        file_toolbar = QToolBar("File")
        file_toolbar.setIconSize(QSize(14, 14))
        self.addToolBar(file_toolbar)

        new_xpa_action = QAction(QIcon(os.path.join('images', 'new_file.png')), "새 작업", self)
        new_xpa_action.setShortcut('Ctrl+N')
        new_xpa_action.setStatusTip("새 작업을 만듭니다.")
        new_xpa_action.triggered.connect(self.new_xpa)
        file_menu.addAction(new_xpa_action)
        file_menu.addSeparator()
        file_toolbar.addAction(new_xpa_action)

        save_xpa_action = QAction(QIcon(os.path.join('images', 'disk.png')), "저장", self)
        save_xpa_action.setShortcut('Ctrl+S')
        save_xpa_action.setStatusTip('작업을 저장합니다.')
        save_xpa_action.triggered.connect(self.save_xpa)
        file_menu.addAction(save_xpa_action)
        file_toolbar.addAction(save_xpa_action)

        saveas_xpa_action = QAction(QIcon(os.path.join('images', 'disk--pencil.png')), "다른 이름으로 저장", self)
        saveas_xpa_action.setShortcut('Ctrl+Shift+S')
        saveas_xpa_action.setStatusTip('작업을 다른 이름으로 저장합니다.')
        # saveas_xpa_action.triggered.connect(self.saveas_xpa)
        file_menu.addAction(saveas_xpa_action)
        file_toolbar.addAction(saveas_xpa_action)

        load_xpa_action = QAction(QIcon(os.path.join('images', 'blue-folder-open-document.png')), "불러오기", self)
        load_xpa_action.setShortcut('Ctrl+O')
        load_xpa_action.setStatusTip('작업을 불러옵니다.')
        # load_xpa_action.triggered.connect(self.open_xpa)
        file_menu.addAction(load_xpa_action)
        file_menu.addSeparator()
        file_toolbar.addAction(load_xpa_action)

        exit_action = QAction(QIcon(os.path.join('images', 'exit.png')), '종료', self)
        exit_action.setStatusTip('프로그램을 종료합니다.')
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)

    def new_xpa(self):
        global NOT_SAVE
        # 새 파일이거나 저장을 했을 경우
        if not NOT_SAVE:
            self._new_xpa_new_dialog()
        else:
            msgBox = QuestionMsgBox('저장하지 않고 새 파일을 만드시겟습니까?', '경고')
            if msgBox.exec_() == QMessageBox.Yes:
                print('확인?')
                self._new_xpa_new_dialog()
            else:
                print('불가!')

    def _new_xpa_new_dialog(self):
        new_dialog = XpaNewActionDialog()
        if new_dialog.exec_():
            filepath, comment = new_dialog.ExcelInfo()
            MakeXpaNew(filepath, comment)
            print(new_dialog.ExcelInfo())
            self._MakeTabList()
        else:
            print('취소')

    def save_xpa(self):
        pass

    def _MakeTabList(self):
        save_path = os.path.join(WORK_DIR, "working.xlsx")
        try:
            self.wb = openpyxl.load_workbook(save_path)
        except FileNotFoundError:
            msgBox = CriticalMsgBox('엑셀 파일 열기에 실패했습니다.')
            msgBox.exec_()
        else:
            sheet_name_list = self.wb.sheetnames
            if '표지' in sheet_name_list:
                sheet_name_list.remove('표지')
            for sheet_name in sheet_name_list:
                sheet = self.wb[sheet_name]
                row = (sheet.max_row - 1) // 19
                new_tab_bar = MyTabBarList()
                new_tab_bar.TableWidget.setRowCount(row)
                self.centralWidget().ui.SheetList.addTab(new_tab_bar, sheet_name)

    #  상태 바 설정
    def MyStatusBar(self):
        status = QStatusBar()
        self.setStatusBar(status)


if __name__ == '__main__':
    app = QApplication(sys.argv)
    myWindow = MainWindow()
    myWindow.show()
    app.exec_()
